from openpyxl import load_workbook


class ContactXls:
    def contact_file(self):
        filename = "contacts.xlsx"
        wb = load_workbook(filename)
        ws = wb.active
        numbers = ws.cell(row=2, column=1)
        print(numbers.value)
