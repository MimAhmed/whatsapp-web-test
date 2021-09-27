from selenium import webdriver
import time
import unittest
import HtmlTestRunner
from Pages.seen_status import SeenStatus
from Locator import locators
from openpyxl import load_workbook

# read contact from excel file
filename = "contacts.xlsx"
wb = load_workbook(filename)
ws = wb.active
numbers = ws.cell(row=4, column=1)


class SearchField(unittest.TestCase):
    driver = None

    @classmethod
    def setUpClass(cls):
        cls.driver = webdriver.Chrome(executable_path="C:/driver/chromedriver_win32_93/chromedriver.exe")
        cls.driver.implicitly_wait(10)
        cls.driver.maximize_window()

    def test_001_seen_status(self):
        driver = self.driver
        driver.get(locators.BASE_URl)
        status = SeenStatus(driver)
        status.is_seen()

    @classmethod
    def tearDownClass(cls):
        cls.driver.close()
        cls.driver.quit()
        print("Test Completed")


if __name__ == '__main__':
    unittest.main(testRunner=HtmlTestRunner.HTMLTestRunner(output='C:/Users/Joy/PycharmProjects/WhatsAppTest/Report'))
