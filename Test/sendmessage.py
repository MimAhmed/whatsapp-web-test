from selenium import webdriver
import time
import unittest
import HtmlTestRunner
from Pages.send_message import SendMessage
from Locator import locators
from openpyxl import load_workbook

# read contact from excel file
filename = "contacts.xlsx"
wb = load_workbook(filename)
ws = wb.active
numbers = ws.cell(row=4, column=1)


class SearchField(unittest.TestCase):
    driver = None

    @classmethod
    def setUpClass(cls):
        cls.driver = webdriver.Chrome(executable_path="C:/driver/chromedriver_win32_93/chromedriver.exe")
        cls.driver.implicitly_wait(10)
        cls.driver.maximize_window()

    def test_001_send_message(self):
        driver = self.driver
        driver.get(locators.BASE_URl)
        number = SendMessage(driver)
        number.click_search_bar()
        number.search_number(numbers.value)  # pass the value from excel sheet
        time.sleep(2)
        number.click_number_filed()
        time.sleep(2)
        number.send_message()
        time.sleep(2)
        ws["B4"] = "sent"
        wb.save(filename)

    @classmethod
    def tearDownClass(cls):
        cls.driver.close()
        cls.driver.quit()
        print("Test Completed")


if __name__ == '__main__':
    unittest.main(testRunner=HtmlTestRunner.HTMLTestRunner(output='C:/Users/Joy/PycharmProjects/WhatsAppTest/Report'))
